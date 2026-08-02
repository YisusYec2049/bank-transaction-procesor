"""
Construye los fixtures de `tests/fixtures/` a partir de archivos reales.

    python scripts/generar_fixtures.py --origen ~/Downloads/bancos21julio

Los archivos reales NO viven en el repo (traen datos de personas). Este script
los lee de donde estén, los pasa por `scripts/anonimizar.py` y escribe fixtures
que sí se pueden comitear.

Sobre los PDF de Bancolombia: el fixture no es un PDF, es un JSON con lo que
`pdfplumber` extrae de él — el texto de cada página y cada palabra con su
posición X. Se hizo así por dos razones. Anonimizar un PDF conservando las
coordenadas exactas obligaría a regenerarlo con otra librería, y esas
coordenadas no son un detalle: `_build_sucursal_lookup()` resuelve la sucursal
por la columna X ∈ [240, 329). Y de paso el fixture deja de depender de la
versión de pdfplumber, que es justo lo que no se quiere estar probando.

Con ese JSON, `parse_pdf()` corre COMPLETO en los tests (el recorte de
sucursal, el pegado de líneas partidas, `_separar_sucursal_desc`,
`_extraer_refs`, los filtros); lo único que se sustituye es la lectura del
archivo.

Fuentes sin fixture todavía, por falta de un archivo de muestra: Colpatria,
Davivienda y PayU (que además necesita dos archivos emparejados).
"""

import argparse
import csv
import io
import json
import logging
import re
import sys
from pathlib import Path

RAIZ = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RAIZ))

import pdfplumber  # noqa: E402

from scripts import anonimizar  # noqa: E402

log = logging.getLogger('fixtures')

DESTINO = RAIZ / 'tests' / 'fixtures'

# Columna X de SUCURSAL en el extracto de Bancolombia, igual que en
# fuentes/bancolombia_2576.py::_build_sucursal_lookup. Una palabra que cae ahí
# es el nombre de una oficina del banco, no de una persona.
_SUCURSAL_X = (240, 329)

# Palabras que forman parte de la estructura del extracto, no de los datos.
_ESTRUCTURALES = {
    'FECHA', 'DESCRIPCIÓN', 'DESCRIPCION', 'SUCURSAL', 'VALOR', 'DOCUMENTO',
    'PÁGINA', 'PAGINA', 'SALDO', 'ANTERIOR', 'ACTUAL', 'TOTAL', 'CUENTA',
    'EXTRACTO', 'BANCOLOMBIA', 'MOVIMIENTOS', 'AHORROS', 'CORRIENTE',
    'NIT', 'CLIENTE', 'PERIODO', 'DEL', 'AL', 'DE', 'LA', 'EL', 'Y', 'SA',
    'S.A.', 'SAS', 'S.A.S.', 'LTDA', 'UNIVERSIDAD',
}


def _vocabulario_tipos() -> set[str]:
    """Palabras de los tipos de transacción reconocidos por los parsers.

    Se leen del propio módulo en vez de copiarlas: si mañana se agrega un tipo
    nuevo a `_FIJAS` (cosa que ya pasó tres veces con las variantes de CONSIG),
    el generador se entera solo y no anonimiza esas palabras por error.
    """
    import fuentes.bancolombia_2576 as bc2576
    import fuentes.bancolombia_2833 as bc2833

    vocab = set(_ESTRUCTURALES)
    for mod in (bc2576, bc2833):
        for frase in list(mod._FIJAS) + list(mod._ABIERTAS) + list(mod.DESCRIPCIONES_ELIMINAR):
            vocab.update(frase.upper().split())
    return vocab


def _es_monto_o_fecha(token: str) -> bool:
    return bool(
        re.fullmatch(r'-?[\d,]+\.\d{2}', token)      # 1,234,567.00
        or re.fullmatch(r'\d{4}/\d{2}/\d{2}', token)  # 2026/07/21
        or re.fullmatch(r'\d{2}/\d{2}/\d{4}', token)
    )


def _mapa_pdf(paginas: list[dict], vocab: set[str]) -> dict[str, str]:
    """Decide, token por token, qué se reemplaza y por qué valor."""
    seguros: set[str] = set()   # palabras de sucursal: se conservan
    for pag in paginas:
        for w in pag['words']:
            if _SUCURSAL_X[0] <= w['x0'] < _SUCURSAL_X[1]:
                seguros.add(w['text'].upper())

    mapa: dict[str, str] = {}
    for pag in paginas:
        for w in pag['words']:
            token = w['text']
            if token in mapa or _es_monto_o_fecha(token):
                continue
            arriba = token.upper()
            if arriba in vocab or arriba in seguros:
                continue
            if re.fullmatch(r'\d{6,}', token):
                mapa[token] = anonimizar.documento(token)
            elif re.fullmatch(r'[A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ.]{2,}', arriba):
                # Palabra en mayúsculas que no es tipo de transacción ni
                # sucursal: en este extracto eso es el nombre de una persona
                # (los NEQUI llegan así, sin número).
                mapa[token] = anonimizar.nombre(token)
    return mapa


def _aplicar(texto: str, mapa: dict[str, str]) -> str:
    if not mapa or not texto:
        return texto
    patron = re.compile('|'.join(re.escape(k) for k in sorted(mapa, key=len, reverse=True)))
    return patron.sub(lambda m: mapa[m.group(0)], texto)


def pdf_bancolombia(origen: Path, nombre_salida: str) -> Path:
    """PDF real → JSON anonimizado con texto y palabras posicionadas."""
    vocab = _vocabulario_tipos()
    paginas = []
    with pdfplumber.open(origen) as pdf:
        for page in pdf.pages:
            paginas.append({
                'text': page.extract_text() or '',
                'words': [
                    {'text': w['text'], 'x0': round(w['x0'], 2), 'top': round(w['top'], 2)}
                    for w in page.extract_words()
                ],
            })

    mapa = _mapa_pdf(paginas, vocab)
    _REEMPLAZADOS.setdefault(nombre_salida, {}).update(
        {k: v for k, v in mapa.items() if len(k) >= 5})
    for pag in paginas:
        pag['text'] = _aplicar(pag['text'], mapa)
        for w in pag['words']:
            w['text'] = mapa.get(w['text'], w['text'])

    salida = DESTINO / nombre_salida
    salida.parent.mkdir(parents=True, exist_ok=True)
    salida.write_text(json.dumps(paginas, ensure_ascii=False, indent=1), encoding='utf-8')
    log.info('%s: %d páginas, %d tokens anonimizados', nombre_salida, len(paginas), len(mapa))
    return salida


# Cómo tratar cada columna, por nombre de encabezado (en minúsculas).
_COLUMNAS = {
    'documento del pagador':  anonimizar.documento,
    'nombre del pagador':     anonimizar.nombre,
    'email del pagador':      anonimizar.email,
    'telefono del pagador':   anonimizar.telefono,
    # `ref. 2` es el "Programa y Detalle Pago" que teclea quien paga — texto
    # libre donde puede caber cualquier cosa. `ref. 1` NO se toca: es un
    # identificador interno de la pasarela, no un dato de la persona, y
    # cambiarlo rompería su correspondencia con el id de la transacción.
    'ref. 2':                 anonimizar.texto_libre,
    'card name':              anonimizar.nombre,
    'customer email':         anonimizar.email,
    'customer phone':         anonimizar.telefono,
    'customer description':   anonimizar.nombre,
    'shipping name':          anonimizar.nombre,
    'description':            anonimizar.texto_libre,
    'statement descriptor':   anonimizar.texto_libre,
    'card address line1':     anonimizar.texto_libre,
    'card address line2':     anonimizar.texto_libre,
    'card address zip':       anonimizar.documento,
}
# Los custom fields de Stripe: [1] nombre del estudiante, [2] su teléfono,
# [3] el nombre del diplomado. El encabezado real termina en " Value" (la
# columna " Key" trae el identificador del campo, que no es dato de nadie).
_STRIPE_CUSTOM = {
    'checkout custom field 1 value': anonimizar.nombre,
    'checkout custom field 2 value': anonimizar.telefono,
    'checkout custom field 3 value': anonimizar.texto_libre,
}


def csv_generico(origen: Path, nombre_salida: str, limite: int | None = None) -> Path:
    _ACTUAL[0] = nombre_salida
    reglas = {**_COLUMNAS, **_STRIPE_CUSTOM}
    texto = origen.read_text(encoding='utf-8', errors='replace')
    lector = csv.DictReader(io.StringIO(texto))
    campos = lector.fieldnames or []

    filas = []
    for i, fila in enumerate(lector):
        if limite is not None and i >= limite:
            break
        nueva = {}
        for col, val in fila.items():
            regla = reglas.get((col or '').strip().lower())
            nueva[col] = _red_de_seguridad(_aplicar_regla(regla, val) if regla else val)
        filas.append(nueva)

    # Segunda pasada, con el mapa ya completo: el nombre de alguien puede estar
    # en una fila anterior a aquella donde reaparece fuera de su columna.
    filas = [{k: _barrer(v, nombre_salida) for k, v in f.items()} for f in filas]

    salida = DESTINO / nombre_salida
    salida.parent.mkdir(parents=True, exist_ok=True)
    with salida.open('w', encoding='utf-8', newline='') as fh:
        escritor = csv.DictWriter(fh, fieldnames=campos)
        escritor.writeheader()
        escritor.writerows(filas)
    log.info('%s: %d filas', nombre_salida, len(filas))
    return salida


# Valores reales que alguna regla reemplazó, POR fixture. Es lo que verifica la
# salida: si un valor que este archivo dijo reemplazar sigue estando en este
# archivo, la regla se aplicó a medias. Ver verificar().
#
# Se lleva por archivo y no en una bolsa común por un caso real: Bancolombia
# liquida los pagos de WOMPI y anota la referencia de la pasarela en su
# extracto, así que el mismo número vive en los dos orígenes. En el PDF es un
# dato a anonimizar; en el CSV de WOMPI es la `matching_key`, la llave de todo
# el cruce, y tiene que sobrevivir intacta. Compararlos entre sí marcaba como
# fuga algo que es correcto en ambos lados.
_REEMPLAZADOS: dict[str, dict[str, str]] = {}
_ACTUAL: list[str] = ['']


def _aplicar_regla(regla, valor):
    """Aplica una regla de anonimización y anota original → reemplazo."""
    original = '' if valor is None else str(valor)
    resultado = regla(original)
    if resultado != original and len(original.strip()) >= 5:
        _REEMPLAZADOS.setdefault(_ACTUAL[0], {})[original.strip()] = str(resultado).strip()
    return resultado


def _barrer(valor, fixture: str):
    """Pasada final: reemplaza, en CUALQUIER celda, los valores que ya se
    anonimizaron en su propia columna.

    Hace falta porque la gente escribe su nombre donde quiera. El campo
    "Programa y Detalle Pago" de WOMPI es texto libre, y junto al nombre del
    diplomado aparece el nombre de quien paga — el mismo que se anonimizó en la
    columna `nombre del pagador` dos celdas más allá. Anonimizar por columna no
    alcanza cuando el dato se repite fuera de su columna.

    Se aplica sobre el valor ya transformado, así que es idempotente y no
    deshace nada de lo anterior.
    """
    mapa = _REEMPLAZADOS.get(fixture)
    if not mapa or not isinstance(valor, str) or len(valor) < 5:
        return valor
    salida = valor
    for original, reemplazo in mapa.items():
        if len(original) >= 5 and original in salida:
            salida = salida.replace(original, reemplazo)
    return salida


def _red_de_seguridad(valor):
    """Última pasada sobre CUALQUIER celda, tenga o no una regla por encabezado.

    Existe por un caso real: el `ReportePagosWompi` trae 62 columnas, y dos de
    ellas (`WH Raw Payload`, `Gateway Response (raw)`) son el JSON crudo que
    devuelve la pasarela — con el nombre, el documento y el correo de la persona
    adentro. Ninguna regla por nombre de columna las iba a atrapar, y la primera
    versión de este script las dejó pasar enteras.

    El pipeline no lee ninguna de esas dos columnas (`read_pagos_wompi_reporte`
    usa 7 de las 62), así que se redactan completas en vez de intentar limpiar
    el JSON por dentro. Lo que no se lee, no hace falta que exista.
    """
    if not isinstance(valor, str):
        return valor
    texto = valor.strip()
    if len(texto) > 120 and texto[:1] in '{[':
        return '{"redactado": "payload crudo, sin uso en el pipeline"}'
    if '@' in texto:
        return anonimizar.texto_libre(valor)
    return valor


def xlsx_generico(origen: Path, nombre_salida: str, limite: int | None = None) -> Path:
    """Excel real → Excel anonimizado. Se decide por encabezado, igual que el
    CSV, pero buscando la fila de encabezados (los archivos del Sistema
    Financiero no siempre arrancan en la fila 1)."""
    _ACTUAL[0] = nombre_salida
    from openpyxl import Workbook, load_workbook

    libro_origen = load_workbook(origen, read_only=True, data_only=True)
    libro = Workbook()
    libro.remove(libro.active)

    for hoja_origen in libro_origen.worksheets:
        hoja = libro.create_sheet(hoja_origen.title[:31])
        filas = list(hoja_origen.iter_rows(values_only=True))
        if not filas:
            continue

        # Fila de encabezados: la primera con 3+ celdas de texto.
        idx_enc = 0
        for i, fila in enumerate(filas[:20]):
            if sum(1 for c in fila if isinstance(c, str) and c.strip()) >= 3:
                idx_enc = i
                break

        encabezados = [str(c).strip().lower() if c else '' for c in filas[idx_enc]]
        reglas_col = {i: _COLUMNAS[h] for i, h in enumerate(encabezados) if h in _COLUMNAS}
        # Encabezados del Sistema Financiero, que no calzan con los de las pasarelas.
        for i, h in enumerate(encabezados):
            if any(p in h for p in ('correo', 'e-mail', 'email')):
                reglas_col[i] = anonimizar.email
            elif any(p in h for p in ('cliente', 'pagador', 'nombre', 'estudiante', 'deudor')):
                reglas_col[i] = anonimizar.nombre
            elif any(p in h for p in ('telefono', 'teléfono', 'celular')):
                reglas_col[i] = anonimizar.telefono
            elif any(p in h for p in ('documento', 'cedula', 'cédula', 'cruceacces', 'numero_id',
                                      'numero id', 'identificacion', 'identificación')):
                reglas_col[i] = anonimizar.documento

        tope = idx_enc + 1 + limite if limite else len(filas)
        for fila in filas[:idx_enc + 1]:
            hoja.append(list(fila))

        transformadas = []
        for fila in filas[idx_enc + 1:tope]:
            nueva = []
            for i, celda in enumerate(fila):
                regla = reglas_col.get(i)
                valor = (_aplicar_regla(regla, celda)
                         if regla and celda is not None else celda)
                nueva.append(_red_de_seguridad(valor))
            transformadas.append(nueva)

        # Segunda pasada, con el mapa ya completo (ver _barrer).
        for fila in transformadas:
            hoja.append([_barrer(c, nombre_salida) for c in fila])

    libro_origen.close()
    salida = DESTINO / nombre_salida
    salida.parent.mkdir(parents=True, exist_ok=True)
    libro.save(salida)
    log.info('%s: %d hojas', nombre_salida, len(libro.worksheets))
    return salida


# Qué archivo del origen alimenta qué fixture. El patrón es el del nombre real
# que manda cada banco.
#
# Los archivos de cartera del Sistema Financiero (Payu UC, Ingresos PSE y PAYU,
# CARTERA PREVENTIVA) NO están acá todavía, y es a propósito. Traen 40+ columnas
# con encabezados que no se dejan clasificar por nombre: `ID CLIENTE` es un
# número, `DOCUMENTO` guarda un UUID, `llave` mezcla documento y consecutivo.
# La heurística de este script los anonimiza mal en los dos sentidos —deja pasar
# cédulas y a la vez convierte números en nombres—, así que necesitan un mapa
# explícito columna por columna, escrito a partir de lo que lee de verdad
# `utils/excel_cartera.py`. Tarea aparte.
_RECETAS = [
    ('ZIP_1686934*.pdf',          'bc2576_extracto.json',    pdf_bancolombia, None),
    ('ZIP_1910000*.pdf',          'bc2833_extracto.json',    pdf_bancolombia, None),
    ('*report-rs-*.csv',          'wompi_reporte.csv',       csv_generico,    None),
    ('unified_payments*.csv',     'stripe_pagos.csv',        csv_generico,    None),
    ('transaction_by_user*.xlsx', 'placetopay_pagos.xlsx',   xlsx_generico,   80),
    ('ReportePagosWompi*.xlsx',   'wompi_reporte_pagos.xlsx', xlsx_generico,  80),
]


def _texto_de(ruta: Path) -> str:
    """Todo el contenido legible de un fixture, sea JSON, CSV o XLSX."""
    if ruta.suffix in ('.json', '.csv'):
        return ruta.read_text(encoding='utf-8', errors='replace')
    from openpyxl import load_workbook
    libro = load_workbook(ruta, read_only=True, data_only=True)
    try:
        return ' '.join(
            str(celda)
            for hoja in libro.worksheets
            for fila in hoja.iter_rows(values_only=True)
            for celda in fila
            if celda is not None
        )
    finally:
        libro.close()


def verificar(rutas: list[Path]) -> list[str]:
    """Revisa los fixtures ya escritos buscando datos que no debieron sobrevivir.

    No es un extra: la primera versión de este generador dejó pasar 80 correos
    reales dentro de columnas de payload JSON, y lo único que lo destapó fue
    mirar la salida a mano. Un generador de datos anonimizados que no verifica
    su propia salida no sirve — el error que importa es justo el que no se ve.

    Dos chequeos, ninguno confía en las reglas de arriba:

    1. **Ningún correo fuera de `example.com`.** Atrapa columnas que ninguna
       regla cubrió, que es exactamente cómo se escapó el payload crudo.
    2. **Ningún valor que alguna regla dijo haber reemplazado.** Si `Pagador`
       se anonimizó en una hoja pero el mismo nombre sobrevivió en otra columna
       de otra hoja, sale acá. Se comparan los valores reales registrados en
       `_REEMPLAZADOS`, no "todo número largo": los montos, las fechas y los
       ids de transacción se conservan a propósito y no son de nadie.
    """
    problemas = []

    # Dos exclusiones, las dos por falsos positivos medidos al construir esto:
    #
    # · Un nombre inventado puede coincidir con uno real. El generador produjo
    #   "TATIANA" para una persona y otra se llamaba así de verdad; el nombre
    #   suelto ya no identifica a nadie (perdió el apellido, el documento y la
    #   fila), pero la comparación literal no sabe distinguirlo.
    # · Números de menos de 7 dígitos: los identificadores de comercio (190093)
    #   y los montos redondos caen ahí, y no son de ninguna persona. Un
    #   documento colombiano tiene 7 dígitos o más.
    inventados = {n.upper() for n in anonimizar._NOMBRES + anonimizar._APELLIDOS}

    for ruta in rutas:
        texto = _texto_de(ruta)
        prohibidos = {
            v for v in _REEMPLAZADOS.get(ruta.name, set())
            if len(v) >= 6
            and v.upper() not in inventados
            and not (v.isdigit() and len(v) < 7)
        }

        correos = {
            m for m in re.findall(r'[\w.+-]+@[\w-]+\.[\w.-]+', texto)
            if not m.lower().endswith('example.com')
        }
        if correos:
            problemas.append(
                f'{ruta.name}: {len(correos)} correo(s) reales, ej. {sorted(correos)[:2]}')

        sobrevivientes = {v for v in prohibidos if v in texto}
        if sobrevivientes:
            problemas.append(
                f'{ruta.name}: {len(sobrevivientes)} valor(es) reales que una regla '
                f'debió reemplazar, ej. {sorted(sobrevivientes)[:3]}')

    return problemas


def main() -> int:
    logging.basicConfig(level=logging.INFO, format='%(message)s')
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('--origen', action='append', required=True, type=Path,
                    help='Carpeta con archivos reales (se puede repetir).')
    args = ap.parse_args()

    carpetas = []
    for raiz in args.origen:
        raiz = raiz.expanduser()
        if not raiz.is_dir():
            log.error('No existe: %s', raiz)
            return 1
        carpetas.append(raiz)
        carpetas.extend(p for p in raiz.iterdir() if p.is_dir())

    escritos, faltantes = [], []
    for patron, nombre_salida, fn, limite in _RECETAS:
        origen = next(
            (c for carpeta in carpetas
             for c in sorted(carpeta.glob(patron)) + sorted(carpeta.glob(patron.lower()))),
            None)
        if origen is None:
            faltantes.append((patron, nombre_salida))
            continue
        try:
            ruta = (fn(origen, nombre_salida, limite) if limite is not None
                    else fn(origen, nombre_salida))
            escritos.append(ruta)
        except Exception as exc:  # no abortar el resto por un archivo raro
            log.error('%s: %s', nombre_salida, exc)
            faltantes.append((patron, nombre_salida))

    log.info('\n%d fixtures generados en %s', len(escritos), DESTINO)
    for patron, nombre_salida in faltantes:
        log.warning('  sin origen: %-28s (buscaba %s)', nombre_salida, patron)

    problemas = verificar(escritos)
    if problemas:
        log.error('\n*** FUGA DE DATOS REALES — estos fixtures NO se pueden comitear ***')
        for p in problemas:
            log.error('  %s', p)
        return 2

    log.info('Verificación de fugas: limpio (%d fixtures revisados).', len(escritos))
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
