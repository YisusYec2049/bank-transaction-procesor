"""Lectura de los Excel de referencia para el cruce de cartera
(Payu UC.xlsx / Ingresos PSE y PAYU.xlsx), descargados de Google Drive."""

import datetime
import logging
from typing import BinaryIO

import openpyxl

from utils.parser import parse_valor, valor_str

log = logging.getLogger(__name__)


def reparar_mojibake(s: str) -> str:
    """Repara texto UTF-8 mal decodificado como cp1252 (ej. 'JosÃ©' -> 'José'),
    visto en los nombres de la hoja STRIPE_USA (hallazgo del 16 de julio: sin
    esto ninguna comparación por nombre funciona con tildes o ñ). Se aplica
    en _cell_str para TODAS las hojas — no se había verificado si el problema
    también afecta a otras (WOMPI, BC2576, Inscrip, Cartera Preventiva), y el
    chequeo ('Ã'/'Â' en el texto) no cuesta nada cuando el texto ya viene
    limpio."""
    if 'Ã' in s or 'Â' in s:
        try:
            return s.encode('cp1252').decode('utf-8')
        except (UnicodeEncodeError, UnicodeDecodeError):
            pass
    return s


def _cell_str(v) -> str:
    if v is None:
        return ''
    if isinstance(v, float):
        return valor_str(v)
    if isinstance(v, int):
        return str(v)
    return reparar_mojibake(str(v).strip())


def _cell_float(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return parse_valor(str(v))


def _cell_int(v) -> int | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    try:
        return int(s)
    except ValueError:
        return None


def _cell_date(v) -> str | None:
    """Normaliza una celda de fecha a ISO (YYYY-MM-DD), o None si no es fecha."""
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    s = str(v).strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _find_header_row(ws, expected: list[str], max_scan: int = 5):
    """Busca, entre las primeras `max_scan` filas, la que contiene todas las
    columnas de `expected` (comparación case-insensitive). Devuelve
    (fila_1based, {header_normalizado: col_idx_0based}).

    Si un encabezado aparece repetido en la misma fila (ej. "convocatoria" o
    "PAGO" duplicados en CARTERA PREVENTIVA), se queda con la PRIMERA
    aparición — las columnas duplicadas de ese archivo son residuales/vacías,
    la real siempre es la más a la izquierda."""
    wanted = {h.strip().lower() for h in expected}
    best_row, best_map, best_hits = None, {}, -1

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1
    ):
        col_map = {}
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            key = str(cell).strip().lower()
            if key:
                col_map.setdefault(key, col_idx)
        hits = len(wanted & col_map.keys())
        if hits > best_hits:
            best_row, best_map, best_hits = row_idx, col_map, hits

    missing = wanted - best_map.keys()
    if missing:
        raise ValueError(f'No se encontraron las columnas {missing} en las primeras {max_scan} filas.')

    return best_row, best_map


def read_inscrip(path: str | BinaryIO) -> list[dict]:
    """Payu UC.xlsx → hoja Inscrip. [{numero_id, id_inscripcion}, ...]."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb['Inscrip']
        header_row, cols = _find_header_row(ws, ['Numero_ID', 'Id_Inscripcion'])

        rows = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            numero_id = _cell_str(row[cols['numero_id']])
            if not numero_id:
                continue
            rows.append({
                'numero_id':      numero_id,
                'id_inscripcion': _cell_str(row[cols['id_inscripcion']]),
            })
        log.info('Inscrip: %d filas leídas.', len(rows))
        return rows
    finally:
        wb.close()


def read_bancolombia_2576(path: str | BinaryIO) -> list[dict]:
    """Ingresos PSE y PAYU.xlsx → hoja BANCOLOMBIA 2576. [{referencia_1, incp, fecha}, ...].

    `fecha` (columna FECHA) se usa en cruzar.py para sugerir, en llaves
    ambiguas, cuál INCP corresponde a un pago nuevo por cadencia mensual
    (ver _sugerir_por_cadencia)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb['BANCOLOMBIA 2576']
        header_row, cols = _find_header_row(ws, ['REFERENCIA 1', 'incp', 'FECHA'])

        rows = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            ref1 = _cell_str(row[cols['referencia 1']])
            if not ref1:
                continue
            rows.append({
                'referencia_1': ref1,
                'incp':         _cell_str(row[cols['incp']]),
                'fecha':        _cell_date(row[cols['fecha']]),
            })
        log.info('BANCOLOMBIA 2576 (Ingresos): %d filas leídas.', len(rows))
        return rows
    finally:
        wb.close()


def read_wompi(path: str | BinaryIO) -> list[dict]:
    """Ingresos PSE y PAYU.xlsx → hoja WOMPI. [{email, inscrip, fecha}, ...]."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb['WOMPI']
        header_row, cols = _find_header_row(ws, ['email', 'INSCRIP', 'Fecha'])

        rows = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            email = _cell_str(row[cols['email']]).lower()
            if not email:
                continue
            rows.append({
                'email':   email,
                'inscrip': _cell_str(row[cols['inscrip']]),
                'fecha':   _cell_date(row[cols['fecha']]),
            })
        log.info('WOMPI (Ingresos): %d filas leídas.', len(rows))
        return rows
    finally:
        wb.close()


def read_stripe_usa(path: str | BinaryIO) -> list[dict]:
    """Ingresos PSE y PAYU.xlsx → hoja STRIPE_USA.
    [{email_cliente, incp, nombre_cliente, fecha}, ...].

    La primera columna de esta hoja no trae un encabezado utilizable (celda A1
    es texto suelto, no un nombre de columna), pero sus datos son siempre la
    fecha/hora del pago — se captura por posición (índice 0), no por nombre.

    `nombre_cliente` (16 de julio, D.2) también se captura por posición
    (índice 4, "NOMBRE CLIENTE") en vez de por header: la hoja tiene una
    SEGUNDA columna llamada "nombre" (índice 16, distinta) que confundiría
    una búsqueda por nombre de columna. Se usa para el cierre de Stripe por
    doble señal (correo + nombre) en cruzar.py."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb['STRIPE_USA']
        header_row, cols = _find_header_row(ws, ['EMAIL CLIENTE', 'INCP'])

        rows = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            email = _cell_str(row[cols['email cliente']]).lower()
            if not email:
                continue
            rows.append({
                'email_cliente':  email,
                'incp':           _cell_str(row[cols['incp']]),
                'nombre_cliente': _cell_str(row[4]) if len(row) > 4 else '',
                'fecha':          _cell_date(row[0]),
            })
        log.info('STRIPE_USA (Ingresos): %d filas leídas.', len(rows))
        return rows
    finally:
        wb.close()


def read_pagos_wompi_reporte(path: str | BinaryIO) -> list[dict]:
    """ReportePagosWompi_*.xlsx (hoja "Pagos Wompi") → pagos automáticos WOMPI
    ("Genera Link") reportados por el Sistema Financiero (Azure). El nombre
    del archivo trae un rango de fechas variable y cambia en cada entrega (se
    busca por patrón en Drive, ver find_latest_file en utils/drive.py).

    Mismo patrón que las demás tablas mirror de "Ingresos PSE y PAYU.xlsx"
    (BANCOLOMBIA 2576 / WOMPI / STRIPE_USA): se reemplaza por completo en
    cada sync (replace_table), no upsert incremental — el archivo de origen
    ya trae acumulado todo lo de entregas anteriores más lo nuevo (confirmado
    con 2 archivos reales del mismo período), así que reemplazar con lo
    último que trae conserva el historial completo igual.

    Algunas entregas traen una segunda hoja ("Hoja1") con un extracto parcial
    de columnas — se ignora a propósito, la fuente real siempre es "Pagos
    Wompi". La fila de encabezado tampoco está siempre en la misma posición
    (algunas entregas traen 2-3 filas de título/período antes), por eso se
    usa _find_header_row con un rango de escaneo más amplio.

    `comprobante` es la llave única del reporte, `documento` es la llave de
    cruce contra `identification` (trae prefijo tipo
    "CC-"/"CEDULA_DE_EXTRANJERIA-", sin normalizar aquí). `inscripcion`,
    `id_transaccion` y `proyecto` (Fase 9, 16 de julio) alimentan el cruce de
    WOMPI LINK en cruzar.py: `inscripcion` (columna "Inscripción", ej.
    "3077") es el número de INCP sin el sufijo del sistema financiero — se
    le busca la forma con sufijo (ej. "3077PN") contra cartera_inscrip antes
    de usarlo; `id_transaccion` (columna "Transaction Id (Wompi)") tiene el
    mismo formato que `id de la transaccion` del CSV de WOMPI y llena
    `cruce_cartera.val`; `proyecto` llena `cruce_cartera.program` (que el
    parser de WOMPI deja vacío a propósito desde Fase 9.5, ver
    fuentes/wompi.py). No se lee "Método Pago" del reporte para
    `metodo_de_pago` — ese campo pasa a ser un literal fijo (9.2), no un dato
    del Excel."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb['Pagos Wompi']
        header_row, cols = _find_header_row(
            ws, ['Comprobante', 'Documento', 'Pagador', 'Fecha Pago',
                 'Inscripción', 'Transaction Id (Wompi)', 'Proyecto'],
            max_scan=10,
        )

        rows = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            comprobante = _cell_str(row[cols['comprobante']])
            if not comprobante:
                continue
            rows.append({
                'comprobante':    comprobante,
                'documento':      _cell_str(row[cols['documento']]),
                'pagador':        _cell_str(row[cols['pagador']]),
                'fecha_pago':     _cell_date(row[cols['fecha pago']]),
                'inscripcion':    _cell_str(row[cols['inscripción']]),
                'id_transaccion': _cell_str(row[cols['transaction id (wompi)']]),
                'proyecto':       _cell_str(row[cols['proyecto']]),
            })
        log.info('ReportePagosWompi (Pagos Wompi): %d filas leídas.', len(rows))
        return rows
    finally:
        wb.close()


def read_cartera_preventiva(path: str | BinaryIO) -> list[dict]:
    """CARTERA PREVENTIVA *.xlsx (hoja Hoja1) → cuotas pendientes por
    inscripción, una fila por cuota. El nombre del archivo trae fecha/versión
    y cambia en cada entrega (se busca por patrón en Drive, ver
    find_latest_file en utils/drive.py).

    Set de columnas confirmado por el usuario el 14 de julio (reemplaza el
    set original del 13 de julio, que incluía convocatoria/tipo
    programa/asesor — se sacaron a propósito, ya no se capturan). El resto de
    columnas del Excel real (DECISIÓN FINAL CARTERA, OBSERVACION, MORA 1,
    ABONO, RETIRO*, REFINANCIACION, etc.) siguen sin usarse.

    `cruce_access` (columna CRUCEACCES) es el documento del deudor — a veces
    trae el NIT con dígito de verificación (ej. "900497967-4"); se guarda tal
    cual, la normalización se hace en el cruce (normalizar_nit).

    `pago` (columna PAGO, aparece duplicada en el Excel real — se toma la
    primera aparición, ver _find_header_row) es un campo de marca/estado del
    proceso manual, se guarda tal cual sin interpretarla."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb['Hoja1']
        header_row, cols = _find_header_row(ws, [
            'llave', 'SITEMA FINANCIERO', 'INSCRIP', 'Cliente', 'MONEDA',
            'correo', 'telefono 1', 'telefono 2', 'F. Vencimiento',
            'DIAS EN CARTERA', 'Valor cuota', 'PAGO', 'Valor a cobrar',
            'Programa', 'CRUCEACCES',
        ])

        rows = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            llave = _cell_str(row[cols['llave']])
            if not llave:
                continue
            rows.append({
                'llave':              llave,
                'sistema_financiero': _cell_str(row[cols['sitema financiero']]),
                'inscrip':            _cell_str(row[cols['inscrip']]),
                'cliente':            _cell_str(row[cols['cliente']]),
                'moneda':             _cell_str(row[cols['moneda']]),
                'correo':             _cell_str(row[cols['correo']]).lower(),
                'telefono_1':         _cell_str(row[cols['telefono 1']]),
                'telefono_2':         _cell_str(row[cols['telefono 2']]),
                'fecha_vencimiento':  _cell_date(row[cols['f. vencimiento']]),
                'dias_en_cartera':    _cell_int(row[cols['dias en cartera']]),
                'valor_cuota':        _cell_float(row[cols['valor cuota']]),
                'pago':               _cell_str(row[cols['pago']]),
                'valor_a_cobrar':     _cell_float(row[cols['valor a cobrar']]),
                'programa':           _cell_str(row[cols['programa']]),
                'cruce_access':       _cell_str(row[cols['cruceacces']]),
            })
        log.info('CARTERA PREVENTIVA: %d filas leídas.', len(rows))
        return rows
    finally:
        wb.close()
